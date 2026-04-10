"""
Excel多表汇总引擎
支持多文件读取、数据合并、智能匹配、汇总计算等功能
"""

import json
import re
import hashlib
from dataclasses import dataclass, field
from datetime import datetime
from typing import Dict, List, Optional, Any, Tuple, Union
from pathlib import Path
from collections import defaultdict


@dataclass
class SheetInfo:
    file_name: str
    sheet_name: str
    headers: List[str]
    row_count: int
    col_count: int
    data_range: str
    has_header: bool = True


@dataclass
class ColumnMapping:
    source_column: str
    target_column: str
    confidence: float = 1.0
    transform_rule: Optional[str] = None


@dataclass
class MergeConfig:
    source_sheets: List[SheetInfo]
    target_columns: List[str]
    column_mappings: List[ColumnMapping]
    merge_mode: str = "append"
    key_columns: List[str] = field(default_factory=list)
    skip_duplicates: bool = True
    fill_na: str = ""


@dataclass
class SummaryResult:
    total_rows: int
    total_columns: int
    source_files: int
    merge_conflicts: int
    duplicate_rows: int
    data: List[Dict[str, Any]]
    statistics: Dict[str, Any]


class ExcelSheetAnalyzer:
    def __init__(self):
        self.supported_extensions = [".xlsx", ".xls", ".csv"]

    def analyze_file(self, file_path: str) -> List[SheetInfo]:
        path = Path(file_path)
        if path.suffix.lower() not in self.supported_extensions:
            return []

        sheets = []
        if path.suffix.lower() == ".csv":
            sheet_info = self._analyze_csv(path)
            if sheet_info:
                sheets.append(sheet_info)
        else:
            sheets = self._analyze_excel(path)

        return sheets

    def _analyze_csv(self, path: Path) -> Optional[SheetInfo]:
        try:
            with open(path, "r", encoding="utf-8-sig") as f:
                lines = f.readlines()

            if not lines:
                return None

            headers = self._parse_csv_line(lines[0])
            row_count = len(lines) - 1
            col_count = len(headers)

            return SheetInfo(
                file_name=path.name,
                sheet_name=path.stem,
                headers=headers,
                row_count=row_count,
                col_count=col_count,
                data_range=f"A1:{self._col_to_letter(col_count)}{row_count + 1}",
            )
        except Exception:
            return None

    def _analyze_excel(self, path: Path) -> List[SheetInfo]:
        sheets = []

        try:
            import openpyxl
            wb = openpyxl.load_workbook(path, read_only=True, data_only=True)

            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]

                headers = []
                row_count = 0
                col_count = 0

                for row in ws.iter_rows(min_row=1, max_row=1):
                    headers = [str(cell.value) if cell.value else f"列{i+1}" for i, cell in enumerate(row)]
                    col_count = len(headers)

                for row in ws.iter_rows(min_row=2):
                    if any(cell.value for cell in row):
                        row_count += 1

                sheets.append(SheetInfo(
                    file_name=path.name,
                    sheet_name=sheet_name,
                    headers=headers,
                    row_count=row_count,
                    col_count=col_count,
                    data_range=f"A1:{self._col_to_letter(col_count)}{row_count + 1}",
                ))

            wb.close()
        except ImportError:
            pass

        return sheets

    def _parse_csv_line(self, line: str) -> List[str]:
        result = []
        current = ""
        in_quotes = False

        for char in line.strip():
            if char == '"':
                in_quotes = not in_quotes
            elif char == "," and not in_quotes:
                result.append(current.strip())
                current = ""
            else:
                current += char

        result.append(current.strip())
        return result

    def _col_to_letter(self, col: int) -> str:
        result = ""
        while col > 0:
            col -= 1
            result = chr(65 + col % 26) + result
            col //= 26
        return result


class ColumnMatcher:
    SIMILARITY_THRESHOLD = 0.6

    COMMON_MAPPINGS = {
        "姓名": ["name", "名字", "员工姓名", "人员姓名", "用户名"],
        "日期": ["date", "时间", "发生日期", "交易日期", "创建时间"],
        "金额": ["amount", "money", "金额", "交易金额", "发生额", "数额"],
        "数量": ["quantity", "count", "数量", "个数", "件数"],
        "单价": ["price", "unit_price", "单价", "价格", "单价金额"],
        "总计": ["total", "sum", "总计", "合计", "总金额", "小计"],
        "部门": ["department", "dept", "部门", "所属部门", "部门名称"],
        "项目": ["project", "项目", "项目名称", "项目名"],
        "备注": ["remark", "note", "备注", "说明", "描述", "注释"],
        "状态": ["status", "state", "状态", "当前状态", "处理状态"],
    }

    def match_columns(
        self,
        source_headers: List[str],
        target_headers: List[str],
    ) -> List[ColumnMapping]:
        mappings = []

        for source_col in source_headers:
            best_match = None
            best_score = 0

            for target_col in target_headers:
                score = self._calculate_similarity(source_col, target_col)

                if score > best_score and score >= self.SIMILARITY_THRESHOLD:
                    best_score = score
                    best_match = target_col

            if best_match:
                mappings.append(ColumnMapping(
                    source_column=source_col,
                    target_column=best_match,
                    confidence=best_score,
                ))

        return mappings

    def _calculate_similarity(self, col1: str, col2: str) -> float:
        c1_lower = col1.lower().strip()
        c2_lower = col2.lower().strip()

        if c1_lower == c2_lower:
            return 1.0

        for standard_name, aliases in self.COMMON_MAPPINGS.items():
            if c1_lower in aliases and c2_lower in aliases:
                return 0.95
            if c1_lower in aliases and c2_lower == standard_name.lower():
                return 0.9
            if c2_lower in aliases and c1_lower == standard_name.lower():
                return 0.9

        if c1_lower in c2_lower or c2_lower in c1_lower:
            return 0.8

        common_chars = set(c1_lower) & set(c2_lower)
        total_chars = set(c1_lower) | set(c2_lower)

        if total_chars:
            jaccard = len(common_chars) / len(total_chars)
            return jaccard

        return 0.0


class MultiSheetMerger:
    def __init__(self):
        self.analyzer = ExcelSheetAnalyzer()
        self.matcher = ColumnMatcher()

    def merge_sheets(
        self,
        file_paths: List[str],
        config: Optional[MergeConfig] = None,
    ) -> SummaryResult:
        all_data = []
        source_files = 0
        merge_conflicts = 0
        duplicate_rows = 0

        if config is None:
            config = self._auto_create_config(file_paths)

        seen_rows = set() if config.skip_duplicates else None

        for file_path in file_paths:
            sheets = self.analyzer.analyze_file(file_path)

            for sheet_info in sheets:
                source_files += 1
                data = self._read_sheet_data(file_path, sheet_info)

                for row in data:
                    mapped_row = self._apply_mapping(row, config.column_mappings)

                    if config.skip_duplicates:
                        row_hash = self._hash_row(mapped_row, config.key_columns)
                        if row_hash in seen_rows:
                            duplicate_rows += 1
                            continue
                        seen_rows.add(row_hash)

                    all_data.append(mapped_row)

        statistics = self._calculate_statistics(all_data, config.target_columns)

        return SummaryResult(
            total_rows=len(all_data),
            total_columns=len(config.target_columns),
            source_files=source_files,
            merge_conflicts=merge_conflicts,
            duplicate_rows=duplicate_rows,
            data=all_data,
            statistics=statistics,
        )

    def _auto_create_config(self, file_paths: List[str]) -> MergeConfig:
        all_headers = set()
        all_sheets = []

        for file_path in file_paths:
            sheets = self.analyzer.analyze_file(file_path)
            all_sheets.extend(sheets)
            for sheet in sheets:
                all_headers.update(sheet.headers)

        target_columns = list(all_headers)
        column_mappings = []

        return MergeConfig(
            source_sheets=all_sheets,
            target_columns=target_columns,
            column_mappings=column_mappings,
        )

    def _read_sheet_data(
        self,
        file_path: str,
        sheet_info: SheetInfo,
    ) -> List[Dict[str, Any]]:
        data = []
        path = Path(file_path)

        if path.suffix.lower() == ".csv":
            data = self._read_csv_data(path, sheet_info)
        else:
            data = self._read_excel_data(path, sheet_info)

        return data

    def _read_csv_data(
        self,
        path: Path,
        sheet_info: SheetInfo,
    ) -> List[Dict[str, Any]]:
        data = []

        try:
            with open(path, "r", encoding="utf-8-sig") as f:
                lines = f.readlines()

            if not lines:
                return data

            headers = self.analyzer._parse_csv_line(lines[0])

            for line in lines[1:]:
                values = self.analyzer._parse_csv_line(line)
                row = dict(zip(headers, values))
                data.append(row)
        except Exception:
            pass

        return data

    def _read_excel_data(
        self,
        path: Path,
        sheet_info: SheetInfo,
    ) -> List[Dict[str, Any]]:
        data = []

        try:
            import openpyxl
            wb = openpyxl.load_workbook(path, data_only=True)
            ws = wb[sheet_info.sheet_name]

            headers = []
            for row in ws.iter_rows(min_row=1, max_row=1):
                headers = [str(cell.value) if cell.value else f"列{i+1}" for i, cell in enumerate(row)]

            for row in ws.iter_rows(min_row=2):
                row_data = {}
                has_data = False

                for i, cell in enumerate(row):
                    if i < len(headers):
                        value = cell.value
                        if value is not None:
                            has_data = True
                        row_data[headers[i]] = value

                if has_data:
                    data.append(row_data)

            wb.close()
        except ImportError:
            pass

        return data

    def _apply_mapping(
        self,
        row: Dict[str, Any],
        mappings: List[ColumnMapping],
    ) -> Dict[str, Any]:
        if not mappings:
            return row

        mapped_row = {}
        for mapping in mappings:
            if mapping.source_column in row:
                value = row[mapping.source_column]
                if mapping.transform_rule:
                    value = self._apply_transform(value, mapping.transform_rule)
                mapped_row[mapping.target_column] = value

        return mapped_row

    def _apply_transform(self, value: Any, rule: str) -> Any:
        if rule == "uppercase":
            return str(value).upper()
        elif rule == "lowercase":
            return str(value).lower()
        elif rule == "strip":
            return str(value).strip()
        return value

    def _hash_row(self, row: Dict[str, Any], key_columns: List[str]) -> str:
        if key_columns:
            key_values = [str(row.get(col, "")) for col in key_columns]
        else:
            key_values = [str(v) for v in row.values()]
        return hashlib.md5("|".join(key_values).encode()).hexdigest()

    def _calculate_statistics(
        self,
        data: List[Dict[str, Any]],
        columns: List[str],
    ) -> Dict[str, Any]:
        stats = {}

        for col in columns:
            values = [row.get(col) for row in data if row.get(col) is not None]

            if not values:
                continue

            numeric_values = []
            for v in values:
                try:
                    numeric_values.append(float(v))
                except (ValueError, TypeError):
                    pass

            col_stats = {
                "count": len(values),
                "non_empty": len([v for v in values if v]),
                "unique": len(set(str(v) for v in values)),
            }

            if numeric_values:
                col_stats.update({
                    "numeric_count": len(numeric_values),
                    "sum": sum(numeric_values),
                    "avg": sum(numeric_values) / len(numeric_values),
                    "min": min(numeric_values),
                    "max": max(numeric_values),
                })

            stats[col] = col_stats

        return stats


class SummaryAggregator:
    def __init__(self):
        self.merger = MultiSheetMerger()

    def create_summary(
        self,
        file_paths: List[str],
        summary_type: str = "general",
    ) -> Dict[str, Any]:
        result = self.merger.merge_sheets(file_paths)

        summary = {
            "overview": {
                "total_records": result.total_rows,
                "total_fields": result.total_columns,
                "source_files": result.source_files,
                "duplicates_removed": result.duplicate_rows,
                "generated_at": datetime.now().isoformat(),
            },
            "field_statistics": result.statistics,
            "data_preview": result.data[:100],
        }

        if summary_type == "financial":
            summary["financial_analysis"] = self._analyze_financial(result.data, result.statistics)
        elif summary_type == "inventory":
            summary["inventory_analysis"] = self._analyze_inventory(result.data, result.statistics)

        return summary

    def _analyze_financial(
        self,
        data: List[Dict[str, Any]],
        stats: Dict[str, Any],
    ) -> Dict[str, Any]:
        analysis = {
            "total_amount": 0,
            "transaction_count": len(data),
            "average_amount": 0,
            "by_category": {},
        }

        amount_keys = ["金额", "amount", "总计", "total", "合计"]
        category_keys = ["类别", "category", "类型", "type", "分类"]

        amount_col = None
        category_col = None

        for key in amount_keys:
            if key in stats:
                amount_col = key
                break

        for key in category_keys:
            for d in data:
                if key in d:
                    category_col = key
                    break
            if category_col:
                break

        if amount_col and amount_col in stats:
            analysis["total_amount"] = stats[amount_col].get("sum", 0)
            analysis["average_amount"] = stats[amount_col].get("avg", 0)

        if category_col and amount_col:
            by_category = defaultdict(float)
            for row in data:
                cat = row.get(category_col, "未分类")
                try:
                    amount = float(row.get(amount_col, 0))
                    by_category[str(cat)] += amount
                except (ValueError, TypeError):
                    pass
            analysis["by_category"] = dict(by_category)

        return analysis

    def _analyze_inventory(
        self,
        data: List[Dict[str, Any]],
        stats: Dict[str, Any],
    ) -> Dict[str, Any]:
        analysis = {
            "total_items": len(data),
            "total_quantity": 0,
            "by_category": {},
        }

        quantity_keys = ["数量", "quantity", "库存", "stock", "件数"]

        for key in quantity_keys:
            if key in stats:
                analysis["total_quantity"] = stats[key].get("sum", 0)
                break

        return analysis


def create_merger() -> MultiSheetMerger:
    return MultiSheetMerger()


def create_aggregator() -> SummaryAggregator:
    return SummaryAggregator()
